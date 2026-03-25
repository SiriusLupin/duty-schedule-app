import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook

# ====== Google Drive API（Service Account）套件 ======
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


# ============================================================
# 0) 使用者可編輯簡化對照表（預設值）
# ============================================================
default_rules = [
    {"原始關鍵字": "調劑複核", "簡化後": "C"},
    {"原始關鍵字": "處方判讀", "簡化後": "判讀"},
    {"原始關鍵字": "藥物諮詢", "簡化後": "諮詢"},
    {"原始關鍵字": "門診藥局調劑", "簡化後": "門診"},
    {"原始關鍵字": "中正 2樓", "簡化後": "中2"},
    {"原始關鍵字": "中正13樓", "簡化後": "中13"},
    {"原始關鍵字": "思源樓", "簡化後": "思源"},
    {"原始關鍵字": "長青樓", "簡化後": "長青"},
    {"原始關鍵字": "抗凝藥師門診", "簡化後": "抗凝門診"},
    {"原始關鍵字": "移植藥師門診", "簡化後": "移植門診"},
    {"原始關鍵字": "中藥局調劑", "簡化後": "中藥局"},
    {"原始關鍵字": "假日非常班之諮詢與藥動服務", "簡化後": "假日oncall"},
    ]


# ============================================================
# 1) Google Drive / Google Sheets API 共用設定
# ============================================================
SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets",
]


def build_credentials():
    """
    從 Streamlit secrets 建立 Service Account 憑證。
    """
    if "gcp_service_account" not in st.secrets:
        st.error("❌ 找不到 st.secrets['gcp_service_account']，請先設定 Streamlit Secrets。")
        st.stop()

    creds = service_account.Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=SCOPES
    )
    return creds


def build_drive_service():
    """建立 Google Drive API client。"""
    creds = build_credentials()
    return build("drive", "v3", credentials=creds)


def build_sheets_service():
    """建立 Google Sheets API client。"""
    creds = build_credentials()
    return build("sheets", "v4", credentials=creds)


# ============================================================
# 2) Google Drive 下載/列檔工具（Service Account）
# ============================================================
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def build_drive_service():
    """
    用 Streamlit secrets 內的 service account 建立 Drive API client。
    你必須先在 Streamlit Cloud 的 Secrets 或 .streamlit/secrets.toml 放入
    [gcp_service_account] 區塊（type/project_id/private_key/client_email/token_uri...）。
    """
    if "gcp_service_account" not in st.secrets:
        st.error("❌ 找不到 st.secrets['gcp_service_account']，請先設定 Streamlit Secrets。")
        st.stop()

    creds = service_account.Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=DRIVE_SCOPES
    )
    return build("drive", "v3", credentials=creds)


def extract_drive_file_id(url: str):
    """
    從使用者貼上的 Google Drive / Google Sheet 連結中抽出 file_id。
    支援常見格式：
    - https://docs.google.com/spreadsheets/d/<ID>/edit...
    - https://drive.google.com/file/d/<ID>/view...
    - https://drive.google.com/open?id=<ID>
    - ...?id=<ID>
    """
    if not url:
        return None

    patterns = [
        r"/d/([a-zA-Z0-9-_]+)",      # .../d/<id>/...
        r"[?&]id=([a-zA-Z0-9-_]+)",  # ...?id=<id> 或 &id=<id>
        r"open\?id=([a-zA-Z0-9-_]+)",
        r"file/d/([a-zA-Z0-9-_]+)",
    ]
    for p in patterns:
        m = re.search(p, url)
        if m:
            return m.group(1)
    return None


def parse_year_month_from_drive_filename(file_name: str):
    """
    解析 Drive 檔名格式：11503班表（民國年3碼 + 月2碼）
    回傳 (year_ad, month, year_month_str) 例如 (2026, 3, "202603")
    抓不到就回 None
    """
    if not file_name:
        return None

    # 支援：11503班表、11503 班表、11503班表.xlsx（若是xlsx也可能有副檔名）
    m = re.search(r"(\d{3})(\d{2})\s*班表", file_name)
    if not m:
        return None

    roc_year = int(m.group(1))        # 例如 115
    month = int(m.group(2))           # 例如 03
    year = roc_year + 1911            # 民國->西元
    year_month = f"{year}{month:02d}" # 例如 202603

    return year, month, year_month


def format_loaded_schedule_name(drive_file_name: str):
    """
    由 Drive 檔名（例如：11503班表）轉成顯示用名稱：115年3月班表
    若解析不到就退回顯示原檔名
    """
    if not drive_file_name:
        return None

    m = re.search(r"(\d{3})(\d{2})\s*班表", drive_file_name)
    if not m:
        return drive_file_name

    roc_year = int(m.group(1))
    month = int(m.group(2))
    return f"{roc_year}年{month}月班表"


def download_drive_file_as_bytes(file_id: str):
    """
    下載 Google Drive 檔案成 BytesIO（記憶體檔案），供 pandas/openpyxl 讀取。
    同時支援：
    A) Google 試算表（原生） -> export 成 xlsx
    B) 真正 .xlsx 檔 -> get_media 直接下載

    回傳：(bio, file_name)
    """
    service = build_drive_service()
    meta = service.files().get(fileId=file_id, fields="name,mimeType").execute()

    file_name = meta.get("name", "")
    mime = meta.get("mimeType", "")

    bio = io.BytesIO()

    if mime == "application/vnd.google-apps.spreadsheet":
        request = service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        request = service.files().get_media(fileId=file_id)

    downloader = MediaIoBaseDownload(bio, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()

    bio.seek(0)
    return bio, file_name


def list_recent_drive_files(months_approx_days: int = 92, page_size: int = 100):
    """
    列出近三個月（約 92 天）內有更新的：
    - Google 試算表
    - Excel .xlsx

    注意：Service Account 只看得到「自己建立」或「別人共享給它」的檔案。
    """
    service = build_drive_service()

    since_dt = datetime.now(timezone.utc) - timedelta(days=months_approx_days)
    since_str = since_dt.isoformat().replace("+00:00", "Z")

    q = (
        "("
        "mimeType='application/vnd.google-apps.spreadsheet' OR "
        "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
        ") "
        f"AND modifiedTime >= '{since_str}' "
        "AND trashed=false"
    )

    resp = service.files().list(
        q=q,
        fields="files(id,name,mimeType,modifiedTime)",
        orderBy="modifiedTime desc",
        pageSize=page_size
    ).execute()

    return resp.get("files", [])


def get_excel_bio(source_choice: str, uploaded_file, selected_drive_file, drive_url_backup: str):
    """
    統一回傳 (BytesIO, drive_file_name)
    - 上傳 Excel：drive_file_name 回 None（因為你說不想用上傳檔名判斷）
    - Drive 下拉選 / 貼連結：drive_file_name 回傳 Drive 檔名
    """
    if source_choice == "上傳 Excel":
        if not uploaded_file:
            return None, None
        data = uploaded_file.read()
        bio = io.BytesIO(data)
        bio.seek(0)
        return bio, None

    if source_choice == "現有共用班表檔案(3個月內)":
        if not selected_drive_file:
            return None, None
        return download_drive_file_as_bytes(selected_drive_file["id"])

    # 貼連結備援（source_choice == "試算表連結"）
    if not drive_url_backup:
        return None, None

    file_id = extract_drive_file_id(drive_url_backup)
    if not file_id:
        st.error("❌ 無法從連結解析檔案 ID，請確認貼的是 Drive/Sheet 分享連結。")
        st.stop()

    try:
        return download_drive_file_as_bytes(file_id)
    except Exception as e:
        st.error(f"❌ 從 Google Drive 下載失敗：{e}")
        st.stop()


# ============================================================
# 3) 灰底假日判斷：第二列日期底色（灰色=假日）
# ============================================================
def build_holiday_map(excel_bio: io.BytesIO) -> dict[int, bool]:
    """
    用 openpyxl 讀取 Excel：
    - 第二列（row=2）日期列的底色（灰底代表假日）
    回傳 holiday_map：{ openpyxl_column_index(1-based): is_holiday }
    """
    excel_bio.seek(0)
    wb = load_workbook(excel_bio, data_only=True)
    ws = wb.active

    # 你目前使用的灰底 RGB（如你的班表底色不同，請改這裡）
    gray_rgb = "FFD9D9D9"

    holiday_map = {}
    for col in range(2, ws.max_column + 1):  # B欄開始（A欄是工作內容）
        cell = ws.cell(row=2, column=col)
        fg = cell.fill.fgColor
        is_gray = (fg.type == "rgb" and fg.rgb == gray_rgb)
        holiday_map[col] = is_gray

    return holiday_map


# ============================================================
# 4) 套用時間規則（含你新增的中2藥局發藥括號時間）
# ============================================================
def apply_time_rules(df, holiday_map, column_map):
    """
    df 欄位應含：日期、星期、工作內容、簡化後內容、Start Time、End Time
    holiday_map：欄位底色假日判定
    column_map： (日期, 星期) -> Excel 欄位 index（B=2 起）
    """
    prescription_time_map = {
        "上午": ("08:00", "12:00"),
        "下午": ("13:30", "17:30"),
        "小夜1hr": ("17:30", "18:30"),
        "小夜": ("17:30", "21:30")
    }

    extra_rows = []

    for idx, row in df.iterrows():
        content = row["工作內容"]
        weekday = str(row["星期"]).strip()

        key = (row["日期"], weekday)
        col_idx = column_map.get(key, None)
        is_holiday = holiday_map.get(col_idx, False)

        # 1) 調劑複核（平日 vs 假日）
        if "調劑複核" in content:
            if is_holiday:
                df.at[idx, "Start Time"] = "11:00"
                df.at[idx, "End Time"] = "15:00"
            else:
                df.at[idx, "Start Time"] = "13:30"
                df.at[idx, "End Time"] = "15:00"

        # 2) 門診藥局調劑（括號時間）
        elif "門診藥局調劑" in content:
            match = re.search(r"\((\d{1,2}:\d{2})-(\d{1,2}:\d{2})\)", content)
            if match:
                df.at[idx, "Start Time"] = match.group(1)
                df.at[idx, "End Time"] = match.group(2)

        # 2.5) 中2藥局發藥（括號時間）
        elif "中2藥局" in content:
            match = re.search(r"\((\d{1,2}:\d{2})-(\d{1,2}:\d{2})\)", content)
            if match:
                df.at[idx, "Start Time"] = match.group(1)
                df.at[idx, "End Time"] = match.group(2)

        # 3) 處方判讀 / 化療處方判讀 / 藥物諮詢 / PreESRD（依上午/下午/小夜）
        elif any(k in content for k in ["處方判讀", "化療處方判讀", "藥物諮詢", "PreESRD"]):
            for key_word, (start, end) in prescription_time_map.items():
                if key_word in content:
                    df.at[idx, "Start Time"] = start
                    df.at[idx, "End Time"] = end
                    break

        # 4) 抗凝藥師門診：週二上午 / 週三下午
        elif "抗凝藥師門診" in content:
            if weekday == "二":
                df.at[idx, "Start Time"] = "08:30"
                df.at[idx, "End Time"] = "12:00"
            elif weekday == "三":
                df.at[idx, "Start Time"] = "13:30"
                df.at[idx, "End Time"] = "17:00"

        # 5) 移植藥師門診：目前只有上午（未來可在此新增下午規則）
        elif "移植藥師門診" in content and "上午" in content:
            df.at[idx, "Start Time"] = "08:30"
            df.at[idx, "End Time"] = "12:00"

        # 6) 中藥局調劑：目前固定 08:30-12:00（若要限制週三可再加條件）
        elif "中藥局調劑" in content:
            df.at[idx, "Start Time"] = "08:30"
            df.at[idx, "End Time"] = "12:00"

        # 7) 瑞德西偉審核：08:00-20:00
        elif "瑞德西偉審核" in content:
            df.at[idx, "Start Time"] = "08:00"
            df.at[idx, "End Time"] = "20:00"

        # 8) 平日：若工作為「處方判讀 7-住院」，額外新增「非常班之諮詢與藥動服務」17:30-21:30
        #if "處方判讀 7-住院" in content and not is_holiday:
        #    extra_rows.append({
        #        "日期": row["日期"],
        #        "星期": row["星期"],
        #        "工作內容": "非常班之諮詢與藥動服務",
        #        "簡化後內容": "小夜oncall",  # 後面仍會做簡化 replace
        #        "Start Time": "17:30",
        #        "End Time": "21:30"
        #    })

        # 9) 假日：「非常班之諮詢與藥動服務」三班
        if "假日非常班之諮詢與藥動服務" in content and is_holiday:
            if "上午" in content:
                df.at[idx, "Start Time"] = "08:00"
                df.at[idx, "End Time"] = "12:30"
            elif "下午" in content:
                df.at[idx, "Start Time"] = "12:30"
                df.at[idx, "End Time"] = "17:00"
            elif "晚上" in content:
                df.at[idx, "Start Time"] = "17:00"
                df.at[idx, "End Time"] = "21:00"

    if extra_rows:
        df = pd.concat([df, pd.DataFrame(extra_rows)], ignore_index=True)

    return df

# ============================================================
# 5) 回饋留言板：Google Sheet 作為後端
# ============================================================
def append_feedback_to_sheet(spreadsheet_id: str, values: list):
    """
    新增一列留言到回饋試算表。
    欄位建議：
    A 時間
    B 暱稱
    C 留言內容
    D 班表來源
    E 班表檔名
    F 代號
    """
    service = build_sheets_service()
    body = {"values": [values]}
    service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range="A1",
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body=body
    ).execute()


def read_feedback_from_sheet(spreadsheet_id: str):
    """
    讀取回饋試算表內容並轉成 DataFrame。
    預設第一列為標題列：
    time, name, message, source, file_name, code
    """
    service = build_sheets_service()
    resp = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range="A:F"
    ).execute()

    values = resp.get("values", [])

    if not values:
        return pd.DataFrame(columns=["time", "name", "message", "source", "file_name", "code"])

    if len(values) == 1:
        return pd.DataFrame(columns=values[0])

    header = values[0]
    rows = values[1:]

    normalized_rows = []
    for row in rows:
        row = row + [""] * (len(header) - len(row))
        normalized_rows.append(row[:len(header)])

    df = pd.DataFrame(normalized_rows, columns=header)

    if "time" in df.columns:
        df = df.sort_values(by="time", ascending=False, kind="mergesort")

    return df


# ============================================================
# 6) 轉換核心邏輯：主程式 tab 共用
# ============================================================
def run_convert(code: str, source: str, excel_bytes: bytes, drive_file_name: str, simplify_map: dict):
    """
    將已載入的班表 bytes + 班表代號 + 縮寫表
    轉為 Google Calendar 可匯入的 CSV DataFrame。
    """
    excel_bio = io.BytesIO(excel_bytes)

    excel_bio.seek(0)
    df = pd.read_excel(excel_bio, header=None)

    holiday_map = build_holiday_map(io.BytesIO(excel_bytes))

    if source in ["現有共用班表檔案(3個月內)", "試算表連結"]:
        parsed = parse_year_month_from_drive_filename(drive_file_name)
        if parsed:
            year, month, year_month = parsed
        else:
            st.error(f"❌ 無法從 Drive 檔名解析年月：{drive_file_name}\n請確認檔名格式為 11503班表")
            return None, None, None
    else:
        title = str(df.iat[0, 0])
        m = re.search(r"(\d{2,3})年(\d{1,2})月", title)
        if not m:
            st.error("❌ 無法從首列標題解析年月，請確認格式如『113年4月班表』")
            return None, None, None
        year = int(m.group(1)) + 1911
        month = int(m.group(2))
        year_month = f"{year}{month:02d}"

    dates = df.iloc[1, 1:].tolist()
    weekdays = df.iloc[2, 1:].tolist()

    date_mapping = [
        {"日期": f"{year}-{month:02d}-{int(d):02d}", "星期": weekdays[i]}
        for i, d in enumerate(dates)
        if str(d).strip().isdigit()
    ]

    col_index_map = {
        (entry["日期"], entry["星期"]): i + 2
        for i, entry in enumerate(date_mapping)
    }

    results = []
    for row_idx in range(3, df.shape[0]):
        raw = df.iat[row_idx, 0]
        if pd.isna(raw):
            continue

        content = str(raw).strip()
        if not content:
            continue
        if content.lower() == "nan":
            continue
        if "附　註" in content:
            continue

        for col_idx in range(1, len(date_mapping) + 1):
            cell = df.iat[row_idx, col_idx]
            cell_str = "" if pd.isna(cell) else str(cell)

            if code in cell_str:
                simplified = re.sub(r"\((\d{1,2}:\d{2})-(\d{1,2}:\d{2})\)", "", content)

                for k, v in simplify_map.items():
                    if pd.notna(k) and pd.notna(v):
                        simplified = simplified.replace(str(k), str(v))

                results.append({
                    "日期": date_mapping[col_idx - 1]["日期"],
                    "星期": date_mapping[col_idx - 1]["星期"],
                    "工作內容": content,
                    "簡化後內容": simplified,
                })

    df_result = pd.DataFrame(results)
    if df_result.empty:
        st.warning("找不到符合此代號的班表內容。請確認代號是否正確，或該月未排班。")
        return None, None, None

    df_result["Start Time"] = ""
    df_result["End Time"] = ""
    df_result = apply_time_rules(df_result, holiday_map, col_index_map)

    df_output = df_result.rename(columns={"簡化後內容": "Subject", "日期": "Start Date"})
    df_output["End Date"] = df_output["Start Date"]
    df_output = df_output[["Subject", "Start Date", "Start Time", "End Date", "End Time"]]

    csv_text = df_output.to_csv(index=False, encoding="utf-8-sig")
    return df_output, csv_text, year_month


# ============================================================
# 7) 更新日誌：純文字但較美觀
# ============================================================
CHANGELOG_ITEMS = [
    {
        "date": "2026-03-06",
        "version": "v3.3",
        "title": "新增三個頁籤",
        "content": "頁面改為「主程式 / 留言回饋 / 更新日誌」，減少單頁資訊量，讓主要操作更集中。"
    },
    {
        "date": "2026-03-06",
        "version": "v3.2",
        "title": "新增留言回饋頁",
        "content": "留言回饋改用 Google Sheet 做後端，可保存時間、暱稱、留言內容、班表來源、班表檔名與代號。"
    },
    {
        "date": "2026-03-06",
        "version": "v3.1",
        "title": "共用班表預設選取優化",
        "content": "預設班表來源改為「現有共用班表檔案(3個月內)」，並自動預選檔名月份最大的班表。"
    },
    {
        "date": "2026-03-05",
        "version": "v3.0",
        "title": "班表來源更新",
        "content": "將班表來源改為連線現有共用班表檔案、上傳班表檔案或班表連結。"
    },
    {
        "date": "2025-03-24",
        "version": "v2.1",
        "title": "提供縮寫清單供使用者自行修改",
        "content": "除預設縮寫外，使用者可自行輸入想要的縮寫"
    },
    {
        "date": "2025-03-24",
        "version": "v2.0",
        "title": "修正假日判別邏輯",
        "content": "以班表日期之底色判定是否為假日"
    },
    {
        "date": "2025-03-21",
        "version": "v1.0",
        "title": "班表轉換工具上線",
        "content": "將電子班表下載後，上傳後可產出日曆檔匯入google日曆"
    },
]


# ============================================================
# 8) 頁面設定與 Session State 初始化
# ============================================================
st.set_page_config(page_title="班表轉換工具", page_icon="📆", layout="centered")

if "loaded_excel_bytes" not in st.session_state:
    st.session_state.loaded_excel_bytes = None
if "loaded_drive_file_name" not in st.session_state:
    st.session_state.loaded_drive_file_name = None
if "last_source" not in st.session_state:
    st.session_state.last_source = None
if "last_code" not in st.session_state:
    st.session_state.last_code = None
if "df_output" not in st.session_state:
    st.session_state.df_output = None
if "csv_text" not in st.session_state:
    st.session_state.csv_text = None
if "year_month" not in st.session_state:
    st.session_state.year_month = None
if "edited_rules" not in st.session_state:
    st.session_state.edited_rules = pd.DataFrame(default_rules)


# ============================================================
# 9) 頁面主體：三個頁籤
# ============================================================
st.title("📆 班表轉換工具")

tab_main, tab_feedback, tab_changelog = st.tabs(["主程式", "留言回饋", "更新日誌"])


# ============================================================
# Tab 1：主程式
# ============================================================
with tab_main:
    try:
        with open("操作說明v3.pdf", "rb") as f:
            st.download_button("📘 下載操作說明 PDF", data=f.read(), file_name="操作說明v3.pdf")
    except FileNotFoundError:
        st.caption("（找不到操作說明 PDF 檔案；若在 Streamlit Cloud 請確認已放入 Repo）")

    status_box = st.empty()

    st.subheader("① 先選班表")

    source = st.radio(
        "選擇班表來源：",
        ["上傳 Excel", "現有共用班表檔案(3個月內)", "試算表連結"],
        index=1,
        horizontal=False
    )

    uploaded_file = None
    selected_drive_file = None
    drive_url_backup = ""

    if source == "上傳 Excel":
        uploaded_file = st.file_uploader("請上傳 Excel 班表（.xlsx）")

    elif source == "現有共用班表檔案(3個月內)":
        try:
            files = list_recent_drive_files(months_approx_days=92, page_size=100)
        # 排除留言回饋試算表
            feedback_sheet_id = st.secrets.get("FEEDBACK_SHEET_ID", "").strip()
            if feedback_sheet_id:
                files = [f for f in files if f["id"] != feedback_sheet_id]

        except Exception as e:
            st.error(f"❌ 無法列出 Google Drive 檔案：{e}")
            files = []

        if not files:
            st.warning("目前 Service Account 近3個月內看不到任何 Excel/試算表。請確認：主管有共享檔案給服務帳號，且檔案近期有更新。")
        else:
            def pretty_label(f):
                typ = "Google試算表" if f["mimeType"] == "application/vnd.google-apps.spreadsheet" else "Excel(.xlsx)"
                mt = f.get("modifiedTime", "")
                return f["name"]

            options = {pretty_label(f): f for f in files}
            labels = list(options.keys())

            def ym_key_from_label(label: str):
                m = re.search(r"(\d{3})(\d{2})\s*班表", label)
                if not m:
                    return -1
                return int(m.group(1)) * 100 + int(m.group(2))

            best_label = max(labels, key=ym_key_from_label)
            default_index = labels.index(best_label)

            chosen = st.selectbox(
                "請選擇班表檔案（近3個月更新）：",
                labels,
                index=default_index
            )
            selected_drive_file = options[chosen]

    else:
        drive_url_backup = st.text_input("請貼上 Google Drive / Google 試算表連結（備援）")

    load_clicked = st.button("📥 載入班表", type="primary")

    if not st.session_state.loaded_excel_bytes:
        status_box.warning("請先選擇班表來源，並按「📥 載入班表」。")

    if load_clicked:
        if source == "上傳 Excel" and uploaded_file is None:
            st.error("❌ 請先上傳 Excel 檔案")
            st.stop()

        if source == "現有共用班表檔案(3個月內)" and selected_drive_file is None:
            st.error("❌ 請先從清單選擇一份班表")
            st.stop()

        if source == "試算表連結" and not drive_url_backup.strip():
            st.error("❌ 請先貼上試算表 / Drive 連結")
            st.stop()

        excel_bio, drive_file_name = get_excel_bio(source, uploaded_file, selected_drive_file, drive_url_backup)

        st.session_state.loaded_excel_bytes = excel_bio.getvalue()
        st.session_state.loaded_drive_file_name = drive_file_name
        st.session_state.last_source = source

        st.session_state.df_output = None
        st.session_state.csv_text = None
        st.session_state.year_month = None

        pretty_name = format_loaded_schedule_name(st.session_state.loaded_drive_file_name)
        if pretty_name:
            status_box.success(f"✅ 班表已載入：{pretty_name}（請輸入代號並轉換）")
        else:
            status_box.success("✅ 班表已載入，請輸入代號並轉換")

    st.subheader("② 再輸入班表代號")
    code = st.text_input("班表代號：", value=(st.session_state.last_code or ""))

    st.subheader("③ 轉換並預覽")
    convert_clicked = st.button("🚀 轉換 / 預覽")

    if convert_clicked:
        if not st.session_state.loaded_excel_bytes:
            st.error("❌ 請先在步驟①按「載入班表」")
        elif not code.strip():
            st.error("❌ 請先輸入班表代號")
        else:
            st.session_state.last_code = code.strip()

            df_rules_now = st.session_state.edited_rules
            simplify_map_now = dict(zip(df_rules_now["原始關鍵字"], df_rules_now["簡化後"]))

            df_output, csv_text, year_month = run_convert(
                code=code.strip(),
                source=st.session_state.last_source,
                excel_bytes=st.session_state.loaded_excel_bytes,
                drive_file_name=st.session_state.loaded_drive_file_name,
                simplify_map=simplify_map_now
            )

            if df_output is not None:
                st.session_state.df_output = df_output
                st.session_state.csv_text = csv_text
                st.session_state.year_month = year_month
                status_box.info("✅ 已完成轉換：請先確認下方預覽，若需要可調整縮寫後重新轉換。")

    if st.session_state.df_output is not None:
        st.subheader("📋 內容預覽")
        st.dataframe(st.session_state.df_output, use_container_width=True)

        st.markdown(
            "<p style='color:red; font-size:18px; font-weight:bold;'>⚠ CSV 檔案直接開啟內容可能為亂碼，但不影響匯入，請先確認上方資料無誤後再下載。</p>",
            unsafe_allow_html=True
        )

        with st.expander("🔧 不滿意？在這裡調整縮寫後重新轉換", expanded=False):
            st.markdown(
                "<p style='color:red; font-size:18px; font-weight:bold;'>🗑️⚠ 注意！若留有空行程式可能發生錯誤，請將空行右側方框勾選後，右上角點選刪除。</p>",
                unsafe_allow_html=True
            )

            edited = st.data_editor(
                st.session_state.edited_rules,
                use_container_width=True,
                num_rows="dynamic",
                key="rules_editor"
            )
            st.session_state.edited_rules = edited
            st.info("✏️ 修改縮寫後，請重新按上方「🚀 轉換 / 預覽」更新結果。")

        st.download_button(
            label=f"📥 下載 {st.session_state.year_month}個人班表({st.session_state.last_code}).csv",
            data=st.session_state.csv_text,
            file_name=f"{st.session_state.year_month}個人班表({st.session_state.last_code}).csv",
            mime="text/csv"
        )


# ============================================================
# Tab 2：留言回饋（回饋型）
# ============================================================
with tab_feedback:
    st.subheader("💬 留言回饋")
    st.caption("這裡可留下問題回報、使用心得或功能建議。")

    feedback_sheet_id = st.secrets.get("FEEDBACK_SHEET_ID", "").strip()

    if not feedback_sheet_id:
        st.warning("尚未設定 FEEDBACK_SHEET_ID，請先於 Streamlit Secrets 設定回饋試算表 ID。")
    else:
        col1, col2 = st.columns([1, 1])
        with col1:
            show_count = st.selectbox("顯示筆數", [10, 20, 50], index=1)
        with col2:
            st.write("")
            st.write("")
            refresh = st.button("🔄 重新整理留言")

        try:
            df_fb = read_feedback_from_sheet(feedback_sheet_id)
        except Exception as e:
            st.error(f"❌ 讀取留言板失敗：{e}")
            df_fb = pd.DataFrame(columns=["time", "name", "message", "source", "file_name", "code"])

        if df_fb.empty:
            st.info("目前還沒有留言，歡迎留下第一則意見。")
        else:
            st.markdown("#### 最新留言")
            for _, row in df_fb.head(show_count).iterrows():
                time_text = str(row.get("time", ""))
                name_text = str(row.get("name", "匿名")) or "匿名"
                msg_text = str(row.get("message", ""))
                source_text = str(row.get("source", ""))
                file_name_text = str(row.get("file_name", ""))
                code_text = str(row.get("code", ""))

                meta = " ｜ ".join([x for x in [source_text, file_name_text, code_text] if x])

                st.markdown(
                    f"""
                    <div style="
                        border:1px solid rgba(128,128,128,0.3);
                        border-radius:12px;
                        padding:12px 14px;
                        margin-bottom:10px;
                    ">
                        <div style="font-weight:700; margin-bottom:4px;">{name_text}</div>
                        <div style="opacity:0.7; font-size:12px; margin-bottom:8px;">{time_text}</div>
                        <div style="white-space:pre-wrap; margin-bottom:8px;">{msg_text}</div>
                        <div style="color:#6B7280; font-size:12px;">{meta}</div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

        st.markdown("---")
        st.markdown("#### ✍️ 我要留言")

        with st.form("feedback_form", clear_on_submit=True):
            nickname = st.text_input("暱稱（可留空，留空則顯示匿名）")
            message = st.text_area("留言內容", height=120, placeholder="例如：某個班別時間規則想調整、某個功能很好用、某個地方不懂等…")
            submitted = st.form_submit_button("送出留言")

        if submitted:
            if not message.strip():
                st.warning("請先輸入留言內容。")
            else:
                try:
                    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    source_text = st.session_state.get("last_source", "") or ""
                    file_name_text = st.session_state.get("loaded_drive_file_name", "") or ""
                    code_text = st.session_state.get("last_code", "") or ""

                    append_feedback_to_sheet(
                        feedback_sheet_id,
                        [
                            now,
                            nickname.strip() or "匿名",
                            message.strip(),
                            source_text,
                            file_name_text,
                            code_text
                        ]
                    )
                    st.success("✅ 已送出留言，謝謝你的回饋！")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 送出留言失敗：{e}")


# ============================================================
# Tab 3：更新日誌（純文字但較美觀）
# ============================================================
with tab_changelog:
    st.subheader("📝 更新日誌")
    st.caption("這裡記錄系統功能調整與版本更新。")

    for item in CHANGELOG_ITEMS:
        st.markdown(
            f"""
            <div style="
            <div style="
                border-left:4px solid #3B82F6;
                padding:12px 14px;
                margin-bottom:12px;
                border-radius:8px;
                border:1px solid rgba(128,128,128,0.25);
            ">
                <div style="font-size:12px; opacity:0.7; margin-bottom:4px;">
                    {item["date"]} ｜ {item["version"]}
                </div>
                <div style="font-weight:700; margin-bottom:6px;">
                    {item["title"]}
                </div>
                <div style="white-space:pre-wrap;">
                    {item["content"]}
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )
