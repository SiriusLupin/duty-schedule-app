import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook


# ========= 判斷灰底是否為假日 ==========
def build_holiday_map(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    gray_rgb = "FFD8D8D8"
    holiday_map = {}
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        is_gray = (
            cell.fill.fgColor.type == "rgb" and cell.fill.fgColor.rgb == gray_rgb
        )
        holiday_map[col] = is_gray
    return holiday_map

# ========= 套用時間規則 ==========
def apply_time_rules(df, holiday_map, column_map):
    prescription_time_map = {
        "上午": ("08:00", "12:00"),
        "下午": ("13:30", "17:30"),
        "小夜1hr": ("17:30", "18:30"),
        "小夜": ("17:30", "21:30")
    }

    extra_rows = []

    for idx, row in df.iterrows():
        content = row["工作內容"]
        weekday = str(row["星期"]).strip()
        key = (row["日期"], weekday)
        col_idx = column_map.get(key, None)
        is_holiday = holiday_map.get(col_idx, False)

        # 調劑複核（平日 vs 假日）
        if "調劑複核" in content:
            if is_holiday:
                df.at[idx, "Start Time"] = "11:00"
                df.at[idx, "End Time"] = "15:00"
            else:
                df.at[idx, "Start Time"] = "13:30"
                df.at[idx, "End Time"] = "15:00"

        elif "門診藥局調劑" in content:
            match = re.search(r"\((\d{1,2}:\d{2})-(\d{1,2}:\d{2})\)", content)
            if match:
                df.at[idx, "Start Time"] = match.group(1)
                df.at[idx, "End Time"] = match.group(2)

        elif any(k in content for k in ["處方判讀", "化療處方判讀", "藥物諮詢", "PreESRD"]):
            for key, (start, end) in prescription_time_map.items():
                if key in content:
                    df.at[idx, "Start Time"] = start
                    df.at[idx, "End Time"] = end
                    break

        elif "抗凝藥師門診" in content:
            if weekday == "二":
                df.at[idx, "Start Time"] = "08:30"
                df.at[idx, "End Time"] = "12:00"
            elif weekday == "三":
                df.at[idx, "Start Time"] = "13:30"
                df.at[idx, "End Time"] = "17:00"

        elif "移植藥師門診" in content and "上午" in content:
            df.at[idx, "Start Time"] = "08:30"
            df.at[idx, "End Time"] = "12:00"

        elif "PreESRD" in content and "上午" in content:
            df.at[idx, "Start Time"] = "08:30"
            df.at[idx, "End Time"] = "12:00"

        elif "中藥局調劑" in content:
            df.at[idx, "Start Time"] = "08:30"
            df.at[idx, "End Time"] = "12:00"

        elif "瑞德西偉審核" in content:
            df.at[idx, "Start Time"] = "08:00"
            df.at[idx, "End Time"] = "20:00"

        if "處方判讀 7-住院" in content and not is_holiday:
            extra_rows.append({
                "日期": row["日期"],
                "星期": row["星期"],
                "工作內容": "非常班之諮詢與藥動服務",
                "簡化後內容": "非常班",
                "Start Time": "17:30",
                "End Time": "21:30"
            })

        if "非常班之諮詢與藥動服務" in content and is_holiday:
            if "上午" in content:
                df.at[idx, "Start Time"] = "08:00"
                df.at[idx, "End Time"] = "12:30"
            elif "下午" in content:
                df.at[idx, "Start Time"] = "12:30"
                df.at[idx, "End Time"] = "17:00"
            elif "晚上" in content:
                df.at[idx, "Start Time"] = "17:00"
                df.at[idx, "End Time"] = "21:00"

    if extra_rows:
        df = pd.concat([df, pd.DataFrame(extra_rows)], ignore_index=True)

    return df

# ========= Streamlit App 主體 ==========
st.title("📆 班表轉換工具（支援假日底色與可編輯簡化對照表）")
with open("班表轉換操作說明v2.pdf", "rb") as f:
    st.download_button("📘 下載操作說明 PDF", data=f.read(), file_name="班表轉換操作說明v2.pdf")
code = st.text_input("請輸入班表代號：")
file = st.file_uploader("請上傳 Excel 班表（.xlsx）")

df_rules = pd.DataFrame(default_rules)
edited_rules = st.data_editor(df_rules, use_container_width=True, num_rows="dynamic")
simplify_map = dict(zip(edited_rules["原始關鍵字"], edited_rules["簡化後"]))

if file and code:
    # ========= 使用者可編輯簡化對照表 ==========
    st.subheader("🔧 字詞縮寫對照表")
    st.markdown(
        "<p style='color:black; font-size:16px; font-weight:bold;'>輸入代碼及上傳檔案後，您可以自行修改想要的縮寫，並可由下方表格預覽。</p>也可點選右上角的"+"新增欄位自訂縮寫"</p>,
        unsafe_allow_html=True
        )
    st.markdown(
        "<p style='color:red; font-size:18px; font-weight:bold;'>⚠注意！若留有空行程式可能發生錯誤，請將空行右側方框勾選後，右上角點選刪除。</p>",
        unsafe_allow_html=True
        )
    default_rules = [
        {"原始關鍵字": "調劑複核", "簡化後": "C"},
        {"原始關鍵字": "處方判讀", "簡化後": "判讀"},
        {"原始關鍵字": "化療處方判讀", "簡化後": "化療判讀"},
        {"原始關鍵字": "藥物諮詢", "簡化後": "諮詢"},
        {"原始關鍵字": "門診藥局調劑", "簡化後": "門診"},
        {"原始關鍵字": "抗凝藥師門診", "簡化後": "抗凝門診"},
        {"原始關鍵字": "移植藥師門診", "簡化後": "移植門診"},
        {"原始關鍵字": "中藥局調劑", "簡化後": "中藥局"},

        ]
    df = pd.read_excel(file, header=None)
    file.seek(0)
    holiday_map = build_holiday_map(file)

    title = str(df.iat[0, 0])
    match = re.search(r"(\d{2,3})年(\d{1,2})月", title)
    if match:
        year = int(match.group(1)) + 1911
        month = int(match.group(2))
        year_month = f"{year}{month:02d}"

        dates = df.iloc[1, 1:].tolist()
        weekdays = df.iloc[2, 1:].tolist()

        date_mapping = [
            {"日期": f"{year}-{month:02d}-{int(d):02d}", "星期": weekdays[i]}
            for i, d in enumerate(dates) if str(d).strip().isdigit()
        ]
        col_index_map = {
            (entry["日期"], entry["星期"]): i + 2
            for i, entry in enumerate(date_mapping)
        }

        results = []
        for row_idx in range(3, df.shape[0]):
            content = str(df.iat[row_idx, 0]).strip()
            if pd.isna(content) or "附　註" in content or content.lower() == "nan":
                continue
            for col_idx in range(1, len(date_mapping) + 1):
                cell = str(df.iat[row_idx, col_idx])
                if code in cell:
                    #simplified = re.sub(r"\(\d{1,2}:\d{2}-\d{1,2}:\d{2}\)", "", content)
                    #for k, v in simplify_map.items():
                    #    simplified = simplified.replace(k, v)
                    simplified = re.sub(r"\(\d{1,2}:\d{2}-\d{1,2}:\d{2}\)", "", content)
                    for k, v in simplify_map.items():
                        if pd.notna(k) and pd.notna(v):
                            simplified = simplified.replace(str(k), str(v))
                    results.append({
                        "日期": date_mapping[col_idx - 1]["日期"],
                        "星期": date_mapping[col_idx - 1]["星期"],
                        "工作內容": content,
                        "簡化後內容": simplified,
                    })

        df_result = pd.DataFrame(results)
        df_result["Start Time"] = ""
        df_result["End Time"] = ""
        df_result = apply_time_rules(df_result, holiday_map, col_index_map)

        df_output = df_result.rename(columns={"簡化後內容": "Subject", "日期": "Start Date"})
        df_output["End Date"] = df_output["Start Date"]
        df_output = df_output[["Subject", "Start Date", "Start Time", "End Date", "End Time"]]

        csv = df_output.to_csv(index=False, encoding="utf-8-sig")
        st.success("✅ 轉換完成，請下載：")

        st.dataframe(df_output, use_container_width=True)
        st.markdown(
    "<p style='color:red; font-size:18px; font-weight:bold;'>⚠ CSV 檔案直接開啟內容可能為亂碼，但不影響匯入，請先確認上方資料無誤後再點選下方按鈕下載。</p>",
    unsafe_allow_html=True
        )
        st.download_button(
            label=f"📥 下載 {year_month}個人班表({code}).csv",
            data=csv,
            file_name=f"{year_month}個人班表({code}).csv",
            mime="text/csv"
        )
        
    else:
        st.error("❌ 無法擷取年份與月份，請確認標題格式如『113年4月班表』")
